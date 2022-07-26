import tkinter as tk
from tkinter import filedialog  # file dialogs help you open, save files or directories
import pandas as pd
from openpyxl import *
from tkinter.messagebox import showinfo



def open():
    path = filedialog.askopenfilename()  # ask for the file location, locate where your file is placed
    df = pd.read_excel(path)  # read the file path
    print(df)  # print the data in file

def save():
    display = entry.get()
    screen_size = entry1.get()
    camera = entry2.get()
    battery = entry3.get()
    ram = entry4.get()
    storage = entry5.get()
    dual_sim = entry6.get()
    wifi = entry7.get()
    bluetooth = entry8.get()
    flash_light = entry9.get()
    phone_color = entry10.get()
    price_range = entry11.get()
    wb = Workbook()
    ws = wb.active
    ws['B1'] = "Display"
    ws['C1'] = "Screen Size"
    ws['D1'] = "Camera"
    ws['E1'] = "Battery"
    ws['F1'] = "RAM"
    ws['G1'] = "Storage"
    ws['H1'] = "Dual Sim"
    ws['I1'] = "Wi-Fi"
    ws['J1'] = "Bluetooth"
    ws['K1'] = "Flash Light"
    ws['L1'] = "Phone Color"
    ws['M1'] = "Price Range"
    ws['B2'] = display
    ws['C2'] = screen_size
    ws['D2'] = camera
    ws['E2'] = battery
    ws['F2'] = ram
    ws['G2'] = storage
    ws['H2'] = dual_sim
    ws['I2'] = wifi
    ws['J2'] = bluetooth
    ws['K2'] = flash_light
    ws['L2'] = phone_color
    ws['M2'] = price_range
    wb.save(r'D:\NIKHIL\Amity University Noida\TCS-ion\Features1.xlsx')  # use Features1.xlsx to store new data
    showinfo("Saved", "Your Entry has been saved")
    file1 = pd.read_excel(
        r'D:\NIKHIL\Amity University Noida\TCS-ion\Features.xlsx')  # use Features.xlsx to store all the data to avoid rewriting
    file2 = pd.read_excel(r'D:\NIKHIL\Amity University Noida\TCS-ion\Features1.xlsx')
    all = [file1, file2]  # define the sequence of data to be stored
    append = pd.concat(all)  # concatenate the files
    append.to_excel(r'D:\NIKHIL\Amity University Noida\TCS-ion\Features.xlsx', index=False)  # write the result to Features.xlsx


def clear():
    entry.delete(0, tk.END)
    entry1.delete(0, tk.END)
    entry2.delete(0, tk.END)
    entry3.delete(0, tk.END)
    entry4.delete(0, tk.END)
    entry5.delete(0, tk.END)
    entry6.delete(0, tk.END)
    entry7.delete(0, tk.END)
    entry8.delete(0, tk.END)
    entry9.delete(0, tk.END)
    entry10.delete(0, tk.END)
    entry11.delete(0, tk.END)

win = tk.Tk()
win.title("Features Requests")
label = tk.Label(win, text="What kind of Hardware you want to feature?", font="lucida 16 bold")
label.grid(row=0, column=1, padx=8, pady=8)
label = tk.Label(win, text="Display")
label.grid(row=1, column=0, padx=8, pady=8)
entry = tk.Entry(win)
entry.grid(row=1, column=1, padx=8, pady=8)
label1 = tk.Label(win, text="Screen Size")
label1.grid(row=2, column=0, padx=8, pady=8)
entry1 = tk.Entry(win)
entry1.grid(row=2, column=1, padx=8, pady=8)
label2 = tk.Label(win, text="Camera")
label2.grid(row=3, column=0, padx=8, pady=8)
entry2 = tk.Entry(win)
entry2.grid(row=3, column=1, padx=8, pady=9)
label3 = tk.Label(win, text="Battery")
label3.grid(row=4, column=0, padx=8, pady=8)
entry3 = tk.Entry(win)
entry3.grid(row=4, column=1, padx=8, pady=8)
label4 = tk.Label(win, text="RAM")
label4.grid(row=5, column=0, padx=8, pady=8)
entry4 = tk.Entry(win)
entry4.grid(row=5, column=1, padx=8, pady=8)
label5 = tk.Label(win, text="Storage")
label5.grid(row=6, column=0, padx=8, pady=8)
entry5 = tk.Entry(win)
entry5.grid(row=6, column=1, padx=8, pady=8)
label6 = tk.Label(win, text="Dual Sim")
label6.grid(row=7, column=0, padx=8, pady=8)
entry6 = tk.Entry(win)
entry6.grid(row=7, column=1, padx=8, pady=8)
label7 = tk.Label(win, text="WI- Fi")
label7.grid(row=8, column=0, padx=8, pady=8)
entry7 = tk.Entry(win)
entry7.grid(row=8, column=1, padx=8, pady=8)
label8 = tk.Label(win, text="Bluetooth")
label8.grid(row=9, column=0, padx=8, pady=8)
entry8 = tk.Entry(win)
entry8.grid(row=9, column=1, padx=8, pady=8)
label9 = tk.Label(win, text="Flash Light")
label9.grid(row=10, column=0, padx=8, pady=8)
entry9 = tk.Entry(win)
entry9.grid(row=10, column=1, padx=8, pady=8)
label10 = tk.Label(win, text="Phone Color")
label10.grid(row=11, column=0, padx=8, pady=8)
entry10 = tk.Entry(win)
entry10.grid(row=11, column=1, padx=8, pady=8)
label11 = tk.Label(win, text="Price Range")
label11.grid(row=12, column=0, padx=8, pady=8)
entry11 = tk.Entry(win)
entry11.grid(row=12, column=1, padx=8, pady=8)

button = tk.Button(win, text="Submit", command=save)
button.grid(row=15, column=0, padx=8, pady=8)
button1 = tk.Button(win, text="Clear", command=clear)
button1.grid(row=15, column=1, padx=8, pady=8)
button2 = tk.Button(win, text="Load Excel", command=open)
button2.grid(row=15, column=2, padx=8, pady=8)


win.mainloop()
