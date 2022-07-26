from tkinter import *
from tkinter import filedialog
import pandas as pd
from openpyxl import *
from tkinter.messagebox import showinfo

root = Tk()
root.geometry('500x400')
root.title('Features')
root.iconbitmap(r'Setting.ico')

def nextPage():
    root.destroy()
    import Hardware

def open():
    path = filedialog.askopenfilename()  # ask for the file location, locate where your file is placed
    df = pd.read_excel(path)  # read the file path
    print(df)  # print the data in file


def save():
    import Hardware
    software = Entry().get()
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Software"
    ws['A2'] = var.get
    wb.save(r'D:\NIKHIL\Amity University Noida\TCS-ion\Features1.xlsx')  # use Features1.xlsx to store new data
    showinfo("Software Chosen!", f"You have choose your software {var.get()} for the phone features")
    file1 = pd.read_excel(
        r'D:\NIKHIL\Amity University Noida\TCS-ion\Features.xlsx')  # use Features.xlsx to store all the data to avoid rewriting
    file2 = pd.read_excel(r'D:\NIKHIL\Amity University Noida\TCS-ion\Features1.xlsx')
    all = [file1, file2]  # define the sequence of data to be stored
    append = pd.concat(all)  # concatenate the files
    append.to_excel(r'D:\NIKHIL\Amity University Noida\TCS-ion\Features.xlsx',
                    index=False)  # write the result to Features.xlsx



var = StringVar()
var.set("Radio")

Label(root, text=" What kind of software you want to feature?", font="lucida 16 bold", justify=LEFT, padx=14).pack()
Radiobutton(root, text="Android", padx=14, variable=var, value="Android").pack(anchor="w")
Radiobutton(root, text="iOS", padx=14, variable=var, value="iOS").pack(anchor="w")
Radiobutton(root, text="Windows Phone OS", padx=14, variable=var, value="Windows Phone OS").pack(anchor="w")
Radiobutton(root, text="Symbian", padx=14, variable=var, value="Symbian").pack(anchor="w")

Button(root, text="Save", command=save).pack()

root.mainloop()
