import tkinter as tk
from tkinter import filedialog                 #file dialogs help you open, save files or directories
import pandas as pd
from openpyxl import *
from tkinter.messagebox import showinfo

def open():
    path = filedialog.askopenfilename()        #ask for the file location, locate where your file is placed
    df = pd.read_excel(path)                  #read the file path
    print(df)                                 #print the data in file

def save():
    software = entry.get()
    os = entry1.get()
    hardware = entry2.get()
    display = entry3.get()
    screen_size = entry4.get()
    camera = entry5.get()
    battery = entry6.get()
    ram = entry7.get()
    storage = entry8.get()
    dual_sim = entry9.get()
    wifi = entry10.get()
    bluetooth = entry11.get()
    flash_light = entry12.get()
    phone_color = entry13.get()
    price_range = entry14.get()
    wb = Workbook()
    ws = wb.active
    ws['B1'] = "OS"
    ws['C1'] = "Hardware"
    ws['D1'] = "Display"
    ws['E1'] = "Screen Size"
    ws['F1'] = "Camera"
    ws['G1'] = "Battery"
    ws['H1'] = "RAM"
    ws['I1'] = "Storage"
    ws['J1'] = "Dual Sim"
    ws['K1'] = "Wi-Fi"
    ws['L1'] = "Bluetooth"
    ws['M1'] = "Flash Light"
    ws['N1'] = "Phone Color"
    ws['O1'] = "Price Range"
    ws['A2'] = software
    ws['B2'] = os
    ws['C2'] = hardware
    ws['D2'] = display
    ws['E2'] = screen_size
    ws['F2'] = camera
    ws['G2'] = battery
    ws['H2'] = ram
    ws['I2'] = storage
    ws['J2'] = dual_sim
    ws['K2'] = wifi
    ws['L2'] = bluetooth
    ws['M2'] = flash_light
    ws['N2'] = phone_color
    ws['O2'] = price_range
    wb.save(r'D:\NIKHIL\Amity University Noida\TCS-ion\Features1.xlsx')     #use Features1.xlsx to store new data
    showinfo("Saved","Your Entry has been saved")
    file1 = pd.read_excel(r'D:\NIKHIL\Amity University Noida\TCS-ion\Features.xlsx')          #use Features.xlsx to store all the data to avoid rewriting
    file2 = pd.read_excel(r'D:\NIKHIL\Amity University Noida\TCS-ion\Features1.xlsx')
    all = [file1, file2]                                                    #define the sequence of data to be stored
    append = pd.concat(all)                                                 #concatenate the files
    append.to_excel(r'D:\NIKHIL\Amity University Noida\TCS-ion\Features.xlsx', index=False)    #write the result to Features.xlsx

def clear():
    entry.delete(0, tk.END)
    entry1.delete(0, tk.END)
    entry2.delete(0, tk.END)
    entry3.delete(0, tk.END)
    entry4.delete(0, tk.END)
    entry5.delete(0, tk.END)
    entry6.delete(0, tk.END)
    entry7.delete(0, tk.END)
    entry8.delete(0, tk.END)
    entry9.delete(0, tk.END)
    entry10.delete(0, tk.END)
    entry11.delete(0, tk.END)
    entry12.delete(0, tk.END)
    entry13.delete(0, tk.END)
    entry14.delete(0, tk.END)

win = tk.Tk()
win.title("Features Requests")
label = tk.Label(win, text="Software")
label.grid(row=0, column=0, padx=8, pady=8)
entry = tk.Entry(win)
entry.grid(row=0, column=1, padx=8, pady=8)
label1 = tk.Label(win, text="OS")
label1.grid(row=1, column=0, padx=8, pady=8)
entry1 = tk.Entry(win)
entry1.grid(row=1, column=1, padx=8, pady=8)
label2 = tk.Label(win, text="Hardware")
label2.grid(row=2, column=0, padx=8, pady=8)
entry2 = tk.Entry(win)
entry2.grid(row=2, column=1, padx=8, pady=9)
label3 = tk.Label(win, text="Display")
label3.grid(row=3, column=0, padx=8, pady=8)
entry3 = tk.Entry(win)
entry3.grid(row=3, column=1, padx=8, pady=8)
label4 = tk.Label(win, text="Screen Size")
label4.grid(row=4, column=0, padx=8, pady=8)
entry4 = tk.Entry(win)
entry4.grid(row=4, column=1, padx=8, pady=8)
label5 = tk.Label(win, text="Camera")
label5.grid(row=5, column=0, padx=8, pady=8)
entry5= tk.Entry(win)
entry5.grid(row=5, column=1, padx=8, pady=8)
label6 = tk.Label(win, text="Battery")
label6.grid(row=6, column=0, padx=8, pady=8)
entry6 = tk.Entry(win)
entry6.grid(row=6, column=1, padx=8, pady=8)
label7 = tk.Label(win, text="RAM")
label7.grid(row=7, column=0, padx=8, pady=8)
entry7 = tk.Entry(win)
entry7.grid(row=7, column=1, padx=8, pady=8)
label8 = tk.Label(win, text="Storage")
label8.grid(row=8, column=0, padx=8, pady=8)
entry8 = tk.Entry(win)
entry8.grid(row=8, column=1, padx=8, pady=8)
label9 = tk.Label(win, text="Dual Sim")
label9.grid(row=9, column=0, padx=8, pady=8)
entry9 = tk.Entry(win)
entry9.grid(row=9, column=1, padx=8, pady=8)
label10 = tk.Label(win, text="Wi-Fi")
label10.grid(row=10, column=0, padx=8, pady=8)
entry10 = tk.Entry(win)
entry10.grid(row=10, column=1, padx=8, pady=8)
label11 = tk.Label(win, text="Bluetooth")
label11.grid(row=11, column=0, padx=8, pady=8)
entry11 = tk.Entry(win)
entry11.grid(row=11, column=1, padx=8, pady=8)
label12 = tk.Label(win, text="Flash Light")
label12.grid(row=12, column=0, padx=8, pady=8)
entry12 = tk.Entry(win)
entry12.grid(row=12, column=1, padx=8, pady=8)
label13 = tk.Label(win, text="Phone Color")
label13.grid(row=13, column=0, padx=8, pady=8)
entry13 = tk.Entry(win)
entry13.grid(row=13, column=1, padx=8, pady=8)
label14 = tk.Label(win, text="Price Range")
label14.grid(row=14, column=0, padx=8, pady=8)
entry14 = tk.Entry(win)
entry14.grid(row=14, column=1, padx=8, pady=8)

button = tk.Button(win, text="Submit", command=save)
button.grid(row=15, column=0, padx=8, pady=8)
button1 = tk.Button(win, text="Clear", command=clear)
button1.grid(row=15, column=1, padx=8, pady=8)
button2 = tk.Button(win, text="Load Excel", command=open)
button2.grid(row=15, column=2, padx=8, pady=8)



win.mainloop()